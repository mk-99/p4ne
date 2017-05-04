#!/usr/local/bin/python3

import glob, ipaddress, pprint, re

from openpyxl import *


def classify(s):
    """
    Take parameters from input string
    :param s: input string
    :return: tuple of (IPv4Interface, score) if applicable or None 
    """

    value = re.match(r'^ ip address ([0-9]+[.][0-9]+[.][0-9]+[.][0-9]+) ([0-9]+[.][0-9]+[.][0-9]+[.][0-9]+)', s)
    if value:
        return (ipaddress.IPv4Interface(value.group(1) + '/' + value.group(2)), 0)

    value = re.match(r'^ ip default-gateway ([0-9]+[.][0-9]+[.][0-9]+[.][0-9]+)', s)
    if value:
        return (ipaddress.IPv4Interface(value.group(1)), 10)

    value = re.match(r'^ standby [0-9]+? ip ([0-9]+[.][0-9]+[.][0-9]+[.][0-9]+)', s)
    if value:
        return (ipaddress.IPv4Interface(value.group(1)), 5)

    value = re.match(r'ip route 0\.0\.0\.0 0\.0\.0\.0 ([0-9]+[.][0-9]+[.][0-9]+[.][0-9]+)', s)
    if value:
        return (ipaddress.IPv4Interface(value.group(1)), 10)

    return None

def find_max(l):
    curr_max = l[0][1]
    retval = l[0][0]
    for i in l:
        if i[1] > curr_max:
            curr_max = i[1]
            retval = i[0]

    return retval


if __name__ == '__main__':

    set_of_ip = set()

    for current_file_name in glob.glob("/Users/mk/Seafile/p4ne_training/config_files/*.txt"):
        with open(current_file_name) as f:
            for current_line in f:
                addr = classify(current_line)
                if addr:
                    set_of_ip.add(addr)

    addr_plan = {}
    hosts = []

    # Fill addr_plan with subnets
    for ip_addr in set_of_ip:
        if ip_addr[0].network.prefixlen != 32: # Subnets
            subnet_key = str(ip_addr[0].network)
            if subnet_key not in addr_plan.keys():
                addr_plan[subnet_key] = [(str(ip_addr[0].ip), ip_addr[1])]
            else:
                addr_plan[subnet_key].append((str(ip_addr[0].ip), ip_addr[1]))
        else: # Hosts
            hosts.append((ip_addr[0].ip, ip_addr[1]))

    # Add gateways to addr_plan
    for subnet_key in addr_plan.keys():
        for gw in hosts:
            if gw[0] in ipaddress.IPv4Network(subnet_key):
                addr_plan[subnet_key].append((str(gw[0]), gw[1]))

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Subnet'
    ws['B1'] = 'Mask'
    ws['C1'] = 'GW'

    current_row = 2

    print("%-20s\t%-4s\t%-20s" % ("Subnet", "Mask", "GW"))

    for subnet_key in addr_plan.keys():
        subnet, mask = subnet_key.split('/')
        mask = '/' + mask
        gw = find_max(addr_plan[subnet_key])

        ws.cell(row=current_row, column=1, value=subnet)
        ws.cell(row=current_row, column=2, value=mask)
        ws.cell(row=current_row, column=3, value=gw)
        current_row += 1

        print("%-20s\t%-4s\t%-20s" % (subnet, mask, gw))

    wb.save('addrplan.xlsx')